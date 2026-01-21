"""
첫콜 마킹 스크립트
- parent_dir 내 child_dir별로 .wav 파일을 파싱
- 고객번호 기준 첫 콜 여부를 판별하여 Excel 파일 생성
"""

from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# 설정
PARENT_DIR = Path("/home/ajung/workspace/sample_call/01")
CHILD_DIRS = ["02", "03", "04", "05", "06", "07"]
OUTPUT_FILE = Path(__file__).parent / "firstcall_report.xlsx"


def parse_filename(filename: str) -> dict:
    """
    파일명 파싱
    포맷: {시분초}-{콜타입}-{대기열 번호}-{고객번호}-{내선번호}-{UniqueID}-{LinkedID}-{RX/TX}.wav
    """
    stem = Path(filename).stem  # 확장자 제거
    parts = stem.rsplit("-", 1)  # RX/TX 분리

    if len(parts) != 2:
        return None

    voice_type = parts[1]  # RX or TX
    main_parts = parts[0].split("-")

    if len(main_parts) < 7:
        return None

    return {
        "time": main_parts[0],           # 시분초 (HHMMSS)
        "call_type": main_parts[1],      # 콜타입 (IN/OUT 등)
        "queue": main_parts[2],          # 대기열 번호
        "customer_number": main_parts[3], # 고객번호
        "extension": main_parts[4],       # 내선번호
        "unique_id": main_parts[5],       # UniqueID
        "linked_id": main_parts[6],       # LinkedID
        "voice_type": voice_type,         # RX(상담사) / TX(고객)
        "call_id": parts[0]               # RX/TX 제외한 전체 (페어 식별용)
    }


def get_unique_calls(day_folder: Path) -> list:
    """
    특정 날짜 폴더에서 유니크한 콜 목록 추출 (RX/TX 페어 중 하나만)
    """
    wav_files = list(day_folder.glob("*.wav"))
    call_ids = set()
    calls = []

    for wav_file in wav_files:
        parsed = parse_filename(wav_file.name)
        if parsed and parsed["call_id"] not in call_ids:
            call_ids.add(parsed["call_id"])
            calls.append({
                "filename": wav_file.name,
                "parsed": parsed
            })

    # 시간순 정렬
    calls.sort(key=lambda x: x["parsed"]["time"])
    return calls


def mark_first_calls(calls: list) -> list:
    """
    고객번호 기준 첫콜 여부 마킹
    """
    seen_customers = set()
    results = []

    for call in calls:
        customer_number = call["parsed"]["customer_number"]
        is_first_call = customer_number not in seen_customers

        if is_first_call:
            seen_customers.add(customer_number)

        results.append({
            "filename": call["parsed"]["call_id"],
            "first_call": "Y" if is_first_call else "N",
            "customer_number": customer_number,
            "time": call["parsed"]["time"]
        })

    return results


def create_excel_report():
    """
    Excel 파일 생성 (각 child_dir별 시트)
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    # 스타일 정의
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    center_align = Alignment(horizontal="center")

    for child_dir in CHILD_DIRS:
        day_folder = PARENT_DIR / child_dir

        if not day_folder.exists():
            print(f"  ⚠ {child_dir} 폴더가 존재하지 않습니다. 건너뜁니다.")
            continue

        # 시트 생성
        ws = wb.create_sheet(title=child_dir)

        # 헤더 설정
        ws["A1"] = "파일명"
        ws["B1"] = "첫콜"
        ws["A1"].font = header_font
        ws["B1"].font = header_font
        ws["A1"].fill = header_fill
        ws["B1"].fill = header_fill
        ws["A1"].alignment = center_align
        ws["B1"].alignment = center_align

        # 데이터 수집 및 첫콜 마킹
        calls = get_unique_calls(day_folder)
        results = mark_first_calls(calls)

        # 데이터 입력
        for idx, result in enumerate(results, start=2):
            ws[f"A{idx}"] = result["filename"]
            ws[f"B{idx}"] = result["first_call"]
            ws[f"B{idx}"].alignment = center_align

        # 컬럼 너비 조정
        ws.column_dimensions["A"].width = 80
        ws.column_dimensions["B"].width = 10

        print(f"  ✓ {child_dir}: {len(results)}건 처리 (첫콜: {sum(1 for r in results if r['first_call'] == 'Y')}건)")

    # 파일 저장
    wb.save(OUTPUT_FILE)
    print(f"\n✓ 결과 파일 생성: {OUTPUT_FILE}")


if __name__ == "__main__":
    print("=" * 60)
    print("첫콜 마킹 스크립트 실행")
    print("=" * 60)
    print(f"Parent Directory: {PARENT_DIR}")
    print(f"Child Directories: {CHILD_DIRS}")
    print("-" * 60)

    create_excel_report()
